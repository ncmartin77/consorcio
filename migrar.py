"""
migrar.py — Migra el Excel existente al esquema actual.

- Agrega hojas nuevas que no existían en versiones anteriores.
- Agrega columnas nuevas a hojas existentes sin tocar los datos.
- Idempotente: se puede ejecutar varias veces sin problema.
- NO modifica ni elimina ningún dato existente.

Uso:
    python migrar.py
    (o lo ejecuta actualizar.bat automáticamente)
"""
import os
import sys
from openpyxl import load_workbook

DB_PATH = os.path.join(os.path.dirname(__file__), "data", "edificio_brasil.xlsx")


def _header(ws):
    """Retorna la lista de encabezados de la fila 1."""
    return [c.value for c in ws[1]]


def ensure_column(ws, col_name, default=""):
    """Agrega col_name al final de la hoja si no existe. Retorna True si se agregó."""
    headers = _header(ws)
    if col_name in headers:
        return False
    col_idx = len(headers) + 1
    ws.cell(row=1, column=col_idx, value=col_name)
    for row in ws.iter_rows(min_row=2):
        ws.cell(row=row[0].row, column=col_idx, value=default)
    return True


def ensure_sheet(wb, name, header):
    """Crea la hoja con su header si no existe. Retorna True si se creó."""
    if name in wb.sheetnames:
        return False
    ws = wb.create_sheet(name)
    ws.append(header)
    return True


def migrate():
    if not os.path.exists(DB_PATH):
        print("INFO: No existe base de datos todavía. Se creará al iniciar la app.")
        return True

    print(f"Migrando: {DB_PATH}")
    wb = load_workbook(DB_PATH)
    changes = []

    # ── Hojas nuevas ──────────────────────────────────────────────────────────

    if ensure_sheet(wb, "GASTOS_RECURRENTES", ["id", "concepto", "categoria"]):
        changes.append("Hoja GASTOS_RECURRENTES creada")

    if ensure_sheet(wb, "LIQUIDACIONES_ESTADO", ["periodo", "estado"]):
        changes.append("Hoja LIQUIDACIONES_ESTADO creada")

    if ensure_sheet(wb, "TAREAS", ["id", "descripcion"]):
        changes.append("Hoja TAREAS creada")

    if ensure_sheet(wb, "PEDIDOS_PRESUPUESTO",
                    ["id", "fecha", "descripcion", "categoria",
                     "estado", "proveedor_elegido", "notas"]):
        changes.append("Hoja PEDIDOS_PRESUPUESTO creada")

    if ensure_sheet(wb, "PRESUPUESTOS",
                    ["id", "pedido_id", "proveedor_id", "proveedor_nombre",
                     "fecha", "importe", "notas", "seleccionado"]):
        changes.append("Hoja PRESUPUESTOS creada")

    # ── Columnas nuevas en hojas existentes ───────────────────────────────────

    if "PROVEEDORES" in wb.sheetnames:
        ws = wb["PROVEEDORES"]
        if ensure_column(ws, "gasto_recurrente", ""):
            changes.append("PROVEEDORES: columna 'gasto_recurrente' agregada")

    if "FACTURAS" in wb.sheetnames:
        ws = wb["FACTURAS"]
        for col in ["numero_factura", "categoria", "extraordinario", "archivo_pdf"]:
            if ensure_column(ws, col, "" if col != "extraordinario" else 0):
                changes.append(f"FACTURAS: columna '{col}' agregada")

    if "UNIDADES" in wb.sheetnames:
        ws = wb["UNIDADES"]
        for col in ["piso", "deuda_inicial"]:
            if ensure_column(ws, col, ""):
                changes.append(f"UNIDADES: columna '{col}' agregada")

    if "CONFIG" in wb.sheetnames:
        ws = wb["CONFIG"]
        config_keys = [r[0].value for r in ws.iter_rows(min_row=1)]
        new_keys = {
            "fecha_simulada": "",
            "dias_cobro": "",
            "horario_cobro": "",
            "direccion_cobro": "",
            "texto_anuncio": "",
            "whatsapp": "",
            "clave_firma": "",
            "url_app": "http://localhost:5000",
        }
        for key, default in new_keys.items():
            if key not in config_keys:
                ws.append([key, default])
                changes.append(f"CONFIG: clave '{key}' agregada")

    # ── Guardar si hubo cambios ───────────────────────────────────────────────

    if changes:
        wb.save(DB_PATH)
        print("Cambios aplicados:")
        for c in changes:
            print(f"  + {c}")
        print(f"\nMigracion completada. {len(changes)} cambio(s) aplicado(s).")
    else:
        print("La base de datos ya está al día. No se requieren cambios.")

    return True


if __name__ == "__main__":
    ok = migrate()
    sys.exit(0 if ok else 1)
