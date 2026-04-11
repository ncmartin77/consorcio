"""
Capa de acceso al Excel (base de datos).
Todas las operaciones de lectura/escritura pasan por aquí.
"""
import os
import secrets
import hashlib
import hmac as _hmac
import calendar as _calendar
from datetime import date, datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
DB_PATH = os.path.join(DATA_DIR, "edificio_brasil.xlsx")
FACTURAS_DIR = os.path.join(DATA_DIR, "facturas")

SHEET_CAJA = "CAJA_DIARIA"
SHEET_GASTOS = "GASTOS_MENSUALES"
SHEET_UNIDADES = "UNIDADES"
SHEET_CATEGORIAS = "CATEGORIAS_PCT"
SHEET_CONFIG = "CONFIG"
SHEET_PROVEEDORES = "PROVEEDORES"
SHEET_FACTURAS = "FACTURAS"
SHEET_PEDIDOS = "PEDIDOS_PRESUPUESTO"
SHEET_PRESUPUESTOS = "PRESUPUESTOS"
SHEET_GASTOS_RECURRENTES = "GASTOS_RECURRENTES"
SHEET_TAREAS = "TAREAS"
LIQPREFIX = "LIQUIDACIONES_"
SHEET_LIQ_ESTADOS = "LIQUIDACIONES_ESTADO"
_LIQ_EST_HEADER = ["periodo", "estado"]


# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------

def _get_wb():
    if not os.path.exists(DB_PATH):
        _init_db()
    return load_workbook(DB_PATH)


def _save_wb(wb):
    wb.save(DB_PATH)


def _init_db():
    os.makedirs(DATA_DIR, exist_ok=True)
    wb = Workbook()
    # Eliminar hoja default
    del wb[wb.sheetnames[0]]

    # CONFIG
    ws = wb.create_sheet(SHEET_CONFIG)
    ws.append(["clave", "valor"])
    ws.append(["edificio_nombre", ""])
    ws.append(["edificio_direccion", ""])
    ws.append(["alias_cbu", ""])
    ws.append(["titular_cuenta", ""])
    ws.append(["administrador", ""])
    ws.append(["telefono", ""])
    ws.append(["email", ""])
    ws.append(["whatsapp", ""])
    ws.append(["tasa_mora", "7"])
    ws.append(["dia_vencimiento", "15"])
    ws.append(["fondo_reserva_mensual", "0"])
    ws.append(["fecha_simulada", ""])
    ws.append(["dias_cobro", ""])
    ws.append(["horario_cobro", ""])
    ws.append(["direccion_cobro", ""])
    ws.append(["texto_anuncio", ""])

    # CATEGORIAS_PCT
    ws = wb.create_sheet(SHEET_CATEGORIAS)
    ws.append(["nombre", "descripcion"])

    # UNIDADES
    ws = wb.create_sheet(SHEET_UNIDADES)
    ws.append(["numero", "descripcion", "propietario", "inquilino", "activo", "piso", "deuda_inicial"])

    # GASTOS_MENSUALES
    ws = wb.create_sheet(SHEET_GASTOS)
    ws.append(["periodo", "concepto", "importe", "tipo"])

    # CAJA_DIARIA
    ws = wb.create_sheet(SHEET_CAJA)
    ws.append(["fecha", "descripcion", "tipo", "categoria", "importe"])

    # PROVEEDORES
    ws = wb.create_sheet(SHEET_PROVEEDORES)
    ws.append(["id", "nombre", "cuit", "telefono", "email", "direccion", "categoria", "notas"])

    # FACTURAS
    ws = wb.create_sheet(SHEET_FACTURAS)
    ws.append(["id", "fecha", "proveedor_id", "proveedor_nombre", "descripcion",
               "importe", "estado", "fecha_pago", "categoria", "numero_factura", "extraordinario", "archivo_pdf"])

    # LIQUIDACIONES_ESTADO
    ws = wb.create_sheet(SHEET_LIQ_ESTADOS)
    ws.append(_LIQ_EST_HEADER)

    # PEDIDOS_PRESUPUESTO
    ws = wb.create_sheet(SHEET_PEDIDOS)
    ws.append(["id", "fecha", "descripcion", "categoria", "estado", "proveedor_elegido", "notas"])

    # PRESUPUESTOS
    ws = wb.create_sheet(SHEET_PRESUPUESTOS)
    ws.append(["id", "pedido_id", "proveedor_id", "proveedor_nombre", "fecha",
               "importe", "notas", "seleccionado"])

    # GASTOS_RECURRENTES
    ws = wb.create_sheet(SHEET_GASTOS_RECURRENTES)
    ws.append(["id", "concepto", "categoria"])

    # TAREAS
    ws = wb.create_sheet(SHEET_TAREAS)
    ws.append(["id", "descripcion"])

    wb.save(DB_PATH)


# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------

def get_config():
    wb = _get_wb()
    ws = wb[SHEET_CONFIG]
    cfg = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            cfg[row[0]] = row[1] or ""
    return cfg


def save_config(data: dict):
    wb = _get_wb()
    ws = wb[SHEET_CONFIG]
    keys = {row[0]: i + 1 for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)) if row[0]}
    for k, v in data.items():
        if k in keys:
            ws.cell(row=keys[k] + 1, column=2).value = v
        else:
            ws.append([k, v])
    _save_wb(wb)


def get_clave_firma() -> str:
    """Retorna la clave secreta HMAC almacenada en CONFIG. La genera si no existe."""
    cfg = get_config()
    clave = cfg.get("clave_firma", "")
    if clave:
        return clave
    clave = secrets.token_hex(32)
    save_config({"clave_firma": clave})
    return clave


def generar_codigo_verificacion(row: dict, cfg: dict, clave_firma: str) -> str:
    """
    Genera un código HMAC-SHA256 de 32 chars hex a partir de los datos del recibo.
    Sirve para verificar que el recibo no fue alterado.
    """
    campos = "|".join([
        str(row.get("periodo", "")),
        str(row.get("unidad", "")),
        str(row.get("descripcion", "")),
        f"{float(row.get('expensas', 0)):.2f}",
        f"{float(row.get('deuda_anterior', 0)):.2f}",
        f"{float(row.get('interes', 0)):.2f}",
        f"{float(row.get('total_a_pagar', 0)):.2f}",
        str(row.get("tipo_pago", "")),
        f"{float(row.get('monto_pagado', 0)):.2f}",
        str(row.get("fecha_pago", "")),
        str(cfg.get("edificio_nombre", "")),
    ])
    mac = _hmac.new(clave_firma.encode(), campos.encode(), hashlib.sha256)
    return mac.hexdigest()[:32]


# ---------------------------------------------------------------------------
# CATEGORIAS DE PORCENTAJE
# ---------------------------------------------------------------------------

def get_categorias():
    wb = _get_wb()
    ws = wb[SHEET_CATEGORIAS]
    cats = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            cats.append({"nombre": row[0], "descripcion": row[1] or ""})
    return cats


def add_categoria(nombre: str, descripcion: str = ""):
    wb = _get_wb()
    ws = wb[SHEET_CATEGORIAS]
    # Verificar duplicado
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == nombre:
            return False
    ws.append([nombre, descripcion])

    # Agregar columna en UNIDADES
    ws_u = wb[SHEET_UNIDADES]
    col_name = f"pct_{nombre}"
    header = [c.value for c in ws_u[1]]
    if col_name not in header:
        ws_u.cell(row=1, column=len(header) + 1).value = col_name

    _save_wb(wb)
    return True


def delete_categoria(nombre: str):
    wb = _get_wb()
    ws = wb[SHEET_CATEGORIAS]
    rows_to_delete = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == nombre:
            rows_to_delete.append(i)
    for r in reversed(rows_to_delete):
        ws.delete_rows(r)
    _save_wb(wb)


# ---------------------------------------------------------------------------
# UNIDADES
# ---------------------------------------------------------------------------

def _unidades_header(ws):
    return [c.value for c in ws[1]]


def get_unidades(solo_activas=False):
    wb = _get_wb()
    ws = wb[SHEET_UNIDADES]
    header = _unidades_header(ws)
    unidades = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        u = dict(zip(header, row))
        if solo_activas and not u.get("activo", True):
            continue
        unidades.append(u)
    return unidades


def get_unidad(numero):
    for u in get_unidades():
        if str(u["numero"]) == str(numero):
            return u
    return None


def save_unidad(data: dict):
    """Crea o actualiza una unidad."""
    wb = _get_wb()
    ws = wb[SHEET_UNIDADES]
    header = _unidades_header(ws)

    # Asegurar columnas de pct
    for cat in get_categorias():
        col_name = f"pct_{cat['nombre']}"
        if col_name not in header:
            ws.cell(row=1, column=len(header) + 1).value = col_name
            header.append(col_name)

    # Buscar fila existente
    target_row = None
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] is not None and str(row[0]) == str(data.get("numero", "")):
            target_row = i
            break

    row_data = [data.get(col, "") for col in header]
    if target_row:
        for col_i, val in enumerate(row_data, start=1):
            ws.cell(row=target_row, column=col_i).value = val
    else:
        ws.append(row_data)

    _save_wb(wb)


def delete_unidad(numero):
    wb = _get_wb()
    ws = wb[SHEET_UNIDADES]
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] is not None and str(row[0]) == str(numero):
            ws.delete_rows(i)
            break
    _save_wb(wb)


# ---------------------------------------------------------------------------
# GASTOS MENSUALES
# ---------------------------------------------------------------------------

def get_gastos(periodo: str):
    """periodo = 'YYYY-MM'"""
    wb = _get_wb()
    ws = wb[SHEET_GASTOS]
    gastos = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if str(row[0]) == periodo:
            gastos.append({
                "row": i,
                "periodo": row[0],
                "concepto": row[1] or "",
                "importe": float(row[2]) if row[2] else 0.0,
                "tipo": row[3] or "FIJO",
            })
    return gastos


def save_gasto(periodo: str, concepto: str, importe: float, tipo: str, row_num: int = None):
    wb = _get_wb()
    ws = wb[SHEET_GASTOS]
    if row_num:
        # No permitir editar el fondo de reserva ya asentado
        existing_tipo = ws.cell(row=row_num, column=4).value
        if existing_tipo == "FONDO_RESERVA":
            return
        ws.cell(row=row_num, column=1).value = periodo
        ws.cell(row=row_num, column=2).value = concepto
        ws.cell(row=row_num, column=3).value = importe
        ws.cell(row=row_num, column=4).value = tipo
    else:
        ws.append([periodo, concepto, importe, tipo])
    _save_wb(wb)


def delete_gasto(row_num: int):
    wb = _get_wb()
    ws = wb[SHEET_GASTOS]
    # No permitir borrar el fondo de reserva ya asentado
    existing_tipo = ws.cell(row=row_num, column=4).value
    if existing_tipo == "FONDO_RESERVA":
        return
    ws.delete_rows(row_num)
    _save_wb(wb)


def get_total_gastos(periodo: str):
    return sum(g["importe"] for g in get_gastos(periodo) if g.get("tipo") != "VARIABLE_FR")


# ---------------------------------------------------------------------------
# CAJA DIARIA
# ---------------------------------------------------------------------------

def get_caja(periodo: str):
    """periodo = 'YYYY-MM'"""
    wb = _get_wb()
    ws = wb[SHEET_CAJA]
    movs = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:
            continue
        fecha = row[0]
        if hasattr(fecha, "strftime"):
            fecha_str = fecha.strftime("%Y-%m-%d")
        else:
            fecha_str = str(fecha)
        if fecha_str[:7] == periodo:
            movs.append({
                "row": i,
                "fecha": fecha_str,
                "descripcion": row[1] or "",
                "tipo": row[2] or "",
                "categoria": row[3] or "",
                "importe": float(row[4]) if row[4] else 0.0,
            })
    movs.sort(key=lambda x: x["fecha"])
    return movs


def save_movimiento(fecha: str, descripcion: str, tipo: str, categoria: str, importe: float, row_num: int = None):
    wb = _get_wb()
    ws = wb[SHEET_CAJA]
    fecha_dt = datetime.strptime(fecha, "%Y-%m-%d").date() if isinstance(fecha, str) else fecha
    if row_num:
        ws.cell(row=row_num, column=1).value = fecha_dt
        ws.cell(row=row_num, column=2).value = descripcion
        ws.cell(row=row_num, column=3).value = tipo
        ws.cell(row=row_num, column=4).value = categoria
        ws.cell(row=row_num, column=5).value = importe
    else:
        ws.append([fecha_dt, descripcion, tipo, categoria, importe])
    _save_wb(wb)


def delete_movimiento(row_num: int):
    wb = _get_wb()
    ws = wb[SHEET_CAJA]
    # Verificar que el período del movimiento no tenga liquidación CERRADA
    row_vals = ws.cell(row=row_num, column=1).value  # fecha
    if row_vals:
        fecha = row_vals
        if hasattr(fecha, "strftime"):
            periodo_mov = fecha.strftime("%Y-%m")
        else:
            periodo_mov = str(fecha)[:7]
        if liq_esta_cerrada(periodo_mov):
            return False
    ws.delete_rows(row_num)
    _save_wb(wb)
    return True


# ---------------------------------------------------------------------------
# LIQUIDACIONES
# ---------------------------------------------------------------------------

def _liq_sheet_name(year: int):
    return f"{LIQPREFIX}{year}"


_LIQ_HEADER = ["periodo", "unidad", "descripcion", "propietario", "inquilino",
               "pct_aplicado", "expensas", "deuda_anterior", "interes", "total_a_pagar",
               "pagado", "monto_pagado", "tipo_pago", "saldo_pendiente", "fecha_pago"]

def _ensure_liq_sheet(wb, year: int):
    name = _liq_sheet_name(year)
    if name not in wb.sheetnames:
        ws = wb.create_sheet(name)
        ws.append(_LIQ_HEADER)
    else:
        ws = wb[name]
        # Migrar hojas existentes que no tienen las columnas nuevas
        header = [c.value for c in ws[1]]
        for col_name in _LIQ_HEADER:
            if col_name not in header:
                ws.cell(row=1, column=len(header) + 1).value = col_name
                header.append(col_name)
    return wb[name]


def get_liquidacion(periodo: str):
    """Retorna filas de liquidación para el periodo 'YYYY-MM'."""
    year = int(periodo[:4])
    wb = _get_wb()
    ws = _ensure_liq_sheet(wb, year)
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if str(row[0]) == periodo:
            fecha_p = row[14] if len(row) > 14 else None
            if hasattr(fecha_p, "strftime"):
                fecha_p = fecha_p.strftime("%Y-%m-%d")
            rows.append({
                "row": i,
                "periodo": row[0],
                "unidad": row[1],
                "descripcion": row[2] or "",
                "propietario": row[3] or "",
                "inquilino": row[4] or "",
                "pct_aplicado": float(row[5]) if row[5] else 0.0,
                "expensas": float(row[6]) if row[6] else 0.0,
                "deuda_anterior": float(row[7]) if row[7] else 0.0,
                "interes": float(row[8]) if row[8] else 0.0,
                "total_a_pagar": float(row[9]) if row[9] else 0.0,
                "pagado": int(row[10]) if row[10] else 0,
                "monto_pagado": float(row[11]) if len(row) > 11 and row[11] else 0.0,
                "tipo_pago": (row[12] if len(row) > 12 and row[12] else "PENDIENTE"),
                "saldo_pendiente": float(row[13]) if len(row) > 13 and row[13] else 0.0,
                "fecha_pago": fecha_p or "",
            })
    return rows


def _prev_periodo(periodo: str):
    y, m = int(periodo[:4]), int(periodo[5:7])
    if m == 1:
        return f"{y-1}-12"
    return f"{y}-{m-1:02d}"


def _next_periodo(periodo: str):
    y, m = int(periodo[:4]), int(periodo[5:7])
    if m == 12:
        return f"{y+1}-01"
    return f"{y}-{m+1:02d}"


def liquidacion_existe(periodo: str) -> bool:
    """Retorna True si ya existe una liquidación generada para este período."""
    year = int(periodo[:4])
    wb = _get_wb()
    ws = _ensure_liq_sheet(wb, year)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == periodo:
            return True
    return False


def generar_liquidacion(periodo: str):
    """
    Genera o recalcula las filas de liquidación para el periodo.
    - Si está CERRADA: inmutable, retorna None.
    - Si está ABIERTA o no existe: genera/recalcula (preservando pagos ya registrados).
    """
    if liq_esta_cerrada(periodo):
        return None  # Cerrada, no se modifica
    # Liquidación a mes vencido: los gastos son del mes ANTERIOR
    mes_gastos = _prev_periodo(periodo)
    ensure_fondo_reserva_gasto(mes_gastos)
    cfg = get_config()
    tasa_mora = float(cfg.get("tasa_mora", 7)) / 100

    # Determinar "hoy" (respeta fecha_simulada)
    fs = cfg.get("fecha_simulada", "").strip() if cfg.get("fecha_simulada") else ""
    if fs:
        try:
            hoy = datetime.strptime(fs, "%Y-%m-%d").date()
        except ValueError:
            hoy = date.today()
    else:
        hoy = date.today()

    # Calcular fecha de vencimiento del período
    try:
        dia_venc = int(cfg.get("dia_vencimiento", 15) or 15)
        year_p, month_p = int(periodo[:4]), int(periodo[5:7])
        ultimo_dia = _calendar.monthrange(year_p, month_p)[1]
        dia_venc_real = ultimo_dia if dia_venc <= 0 else min(dia_venc, ultimo_dia)
        fecha_vencimiento = date(year_p, month_p, dia_venc_real)
    except Exception:
        fecha_vencimiento = None

    unidades = get_unidades(solo_activas=True)
    gastos = get_gastos(mes_gastos)
    total_gastos = sum(g["importe"] for g in gastos if g.get("tipo") != "VARIABLE_FR")
    categorias = [c["nombre"] for c in get_categorias()]

    prev = _prev_periodo(periodo)
    prev_liq = {str(r["unidad"]): r for r in get_liquidacion(prev)}

    year = int(periodo[:4])
    wb = _get_wb()
    ws = _ensure_liq_sheet(wb, year)

    # Guardar pagos existentes ANTES de borrar las filas
    pagos_previos = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == periodo:
            pagos_previos[str(row[1])] = {
                "monto_pagado": float(row[11]) if len(row) > 11 and row[11] else 0.0,
                "tipo_pago":    row[12] if len(row) > 12 and row[12] else "PENDIENTE",
                "fecha_pago":   row[14] if len(row) > 14 else None,
            }

    # Eliminar filas existentes del periodo
    rows_to_del = [i for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2)
                   if str(row[0]) == periodo]
    for r in reversed(rows_to_del):
        ws.delete_rows(r)

    # Recalcular y agregar filas preservando pagos
    expensas_ya_asignadas = 0.0
    for idx, u in enumerate(unidades):
        numero = str(u["numero"])
        pct_vals = []
        for cat in categorias:
            val = u.get(f"pct_{cat}", 0)
            if val:
                try:
                    pct_vals.append(float(val))
                except (ValueError, TypeError):
                    pass
        pct = pct_vals[0] if len(pct_vals) == 1 else (sum(pct_vals) / len(pct_vals) if pct_vals else 0.0)

        if idx == len(unidades) - 1 and total_gastos > 0:
            # Ajuste contable: la última unidad absorbe la diferencia de redondeo
            expensas = round(total_gastos - expensas_ya_asignadas, 2)
        else:
            expensas = round(total_gastos * pct / 100, 2)
            expensas_ya_asignadas += expensas

        prev_row = prev_liq.get(numero)
        if not prev_row:
            try:
                deuda_anterior = float(u.get("deuda_inicial") or 0)
            except (ValueError, TypeError):
                deuda_anterior = 0.0
        elif prev_row["pagado"] == 1:
            deuda_anterior = 0.0
        elif prev_row.get("tipo_pago") == "PARCIAL":
            deuda_anterior = round(prev_row.get("saldo_pendiente", prev_row["total_a_pagar"]), 2)
        else:
            deuda_anterior = round(prev_row["total_a_pagar"], 2)

        # Restaurar pago previo y recalcular estado
        pago = pagos_previos.get(numero, {})
        monto_pagado = pago.get("monto_pagado", 0.0)
        fecha_pago = pago.get("fecha_pago")

        interes_deuda = round(deuda_anterior * tasa_mora, 2)
        mora_corriente = 0.0
        if fecha_vencimiento and hoy > fecha_vencimiento and monto_pagado == 0:
            mora_corriente = round(expensas * tasa_mora, 2)
        interes = round(interes_deuda + mora_corriente, 2)
        total = round(expensas + deuda_anterior + interes, 2)

        if monto_pagado >= total and monto_pagado > 0:
            tipo_pago = "TOTAL"
            saldo = 0.0
            pagado_flag = 1
        elif monto_pagado > 0:
            tipo_pago = "PARCIAL"
            saldo = round(total - monto_pagado, 2)
            pagado_flag = 0
        else:
            tipo_pago = "PENDIENTE"
            saldo = total
            pagado_flag = 0

        ws.append([
            periodo, numero,
            u.get("descripcion", ""),
            u.get("propietario", ""),
            u.get("inquilino", ""),
            pct, expensas, deuda_anterior, interes, total,
            pagado_flag,
            monto_pagado if monto_pagado > 0 else None,
            tipo_pago if monto_pagado > 0 else None,
            saldo if monto_pagado > 0 else None,
            fecha_pago,
        ])

    _save_wb(wb)
    set_liq_estado(periodo, "ABIERTA")
    return get_liquidacion(periodo)


def marcar_pagado(periodo: str, unidad: str, monto_pagado: float, fecha_pago: str = None):
    """
    Registra un pago (total o parcial) para una unidad en un período.
    - monto_pagado = 0 → deshacer pago (volver a pendiente)
    - monto_pagado >= total_a_pagar → pago total
    - 0 < monto_pagado < total_a_pagar → pago parcial
    """
    if liq_esta_cerrada(periodo):
        return {"tipo": "CERRADA", "saldo": 0}
    if not fecha_pago:
        fecha_pago = date.today().strftime("%Y-%m-%d")
    year = int(periodo[:4])
    wb = _get_wb()
    ws = _ensure_liq_sheet(wb, year)
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if str(row[0]) == periodo and str(row[1]) == str(unidad):
            total_a_pagar = float(row[9]) if row[9] else 0.0
            tipo_actual = row[12] or "PENDIENTE"

            # No se puede modificar un pago total
            if tipo_actual == "TOTAL":
                return {"tipo": "TOTAL", "saldo": 0}

            if monto_pagado <= 0:
                return None

            # Si ya hubo un pago parcial, sumar al existente
            monto_anterior = float(row[11]) if row[11] else 0.0
            total_pagado = round(monto_anterior + monto_pagado, 2)

            saldo = round(total_a_pagar - total_pagado, 2)
            if saldo <= 0:
                tipo_pago = "TOTAL"
                saldo = 0.0
                pagado_flag = 1
            else:
                tipo_pago = "PARCIAL"
                pagado_flag = 0

            try:
                fecha_dt = datetime.strptime(fecha_pago, "%Y-%m-%d").date()
            except (ValueError, TypeError):
                fecha_dt = date.today()

            ws.cell(row=i, column=11).value = pagado_flag
            ws.cell(row=i, column=12).value = total_pagado
            ws.cell(row=i, column=13).value = tipo_pago
            ws.cell(row=i, column=14).value = saldo
            ws.cell(row=i, column=15).value = fecha_dt
            _save_wb(wb)

            # Registrar solo el nuevo monto en caja diaria
            desc = row[4] or row[3] or f"Unidad {unidad}"
            sufijo = "(PARCIAL)" if tipo_pago == "PARCIAL" else ""
            descripcion = f"Expensas {periodo} - {desc} - UF {unidad} {sufijo}".strip()
            save_movimiento(
                fecha=fecha_pago,
                descripcion=descripcion,
                tipo="ENTRADA",
                categoria="Expensas",
                importe=round(monto_pagado, 2),
            )
            return {"tipo": tipo_pago, "saldo": saldo, "monto": monto_pagado, "total_pagado": total_pagado}
    _save_wb(wb)
    return None


def marcar_todos_pagado(periodo: str, fecha_pago: str):
    """Marca como pagadas todas las unidades PENDIENTES de un período."""
    if liq_esta_cerrada(periodo):
        return 0
    liq = get_liquidacion(periodo)
    pendientes = [r for r in liq if r.get("tipo_pago", "PENDIENTE") == "PENDIENTE"]
    for r in pendientes:
        marcar_pagado(periodo, str(r["unidad"]), r["total_a_pagar"], fecha_pago)
    return len(pendientes)


def get_años_con_liquidacion():
    wb = _get_wb()
    años = []
    for name in wb.sheetnames:
        if name.startswith(LIQPREFIX) and name != SHEET_LIQ_ESTADOS:
            try:
                años.append(int(name[len(LIQPREFIX):]))
            except ValueError:
                pass
    return sorted(años)


# ---------------------------------------------------------------------------
# ESTADO DE LIQUIDACIÓN (ABIERTA / CERRADA)
# ---------------------------------------------------------------------------

def get_liq_estado(periodo: str) -> str:
    """Returns 'ABIERTA', 'CERRADA', or '' if not generated."""
    wb = _get_wb()
    ws = _ensure_sheet(wb, SHEET_LIQ_ESTADOS, _LIQ_EST_HEADER)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and str(row[0]) == periodo:
            return row[1] or "ABIERTA"
    return ""


def set_liq_estado(periodo: str, estado: str):
    """Creates or updates the state of a liquidation period."""
    wb = _get_wb()
    ws = _ensure_sheet(wb, SHEET_LIQ_ESTADOS, _LIQ_EST_HEADER)
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] and str(row[0]) == periodo:
            ws.cell(row=i, column=2).value = estado
            _save_wb(wb)
            return
    ws.append([periodo, estado])
    _save_wb(wb)


def liq_esta_cerrada(periodo: str) -> bool:
    return get_liq_estado(periodo) == "CERRADA"


def _ensure_sheet(wb, name, header):
    """Crea la hoja con header si no existe (para DBs antiguas)."""
    if name not in wb.sheetnames:
        ws = wb.create_sheet(name)
        ws.append(header)
    return wb[name]


def _next_id(ws):
    """Retorna el próximo ID numérico para una hoja."""
    max_id = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            try:
                max_id = max(max_id, int(row[0]))
            except (ValueError, TypeError):
                pass
    return max_id + 1


# ---------------------------------------------------------------------------
# PROVEEDORES
# ---------------------------------------------------------------------------

_PROV_HEADER = ["id", "nombre", "cuit", "telefono", "email", "direccion", "categoria", "notas", "gasto_recurrente"]

def get_proveedores():
    wb = _get_wb()
    ws = _ensure_sheet(wb, SHEET_PROVEEDORES, _PROV_HEADER)
    result = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] is not None:
            result.append(dict(zip(_PROV_HEADER, row)) | {"row": i})
    return result


def get_proveedor(pid):
    for p in get_proveedores():
        if str(p["id"]) == str(pid):
            return p
    return None


def save_proveedor(data: dict):
    wb = _get_wb()
    ws = _ensure_sheet(wb, SHEET_PROVEEDORES, _PROV_HEADER)
    pid = data.get("id")
    if pid:
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] is not None and str(row[0]) == str(pid):
                for j, col in enumerate(_PROV_HEADER, start=1):
                    ws.cell(row=i, column=j).value = data.get(col, "")
                _save_wb(wb)
                return int(pid)
    new_id = _next_id(ws)
    data["id"] = new_id
    ws.append([data.get(col, "") for col in _PROV_HEADER])
    _save_wb(wb)
    return new_id


def delete_proveedor(pid):
    wb = _get_wb()
    ws = _ensure_sheet(wb, SHEET_PROVEEDORES, _PROV_HEADER)
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] is not None and str(row[0]) == str(pid):
            ws.delete_rows(i)
            break
    _save_wb(wb)


# ---------------------------------------------------------------------------
# GASTOS RECURRENTES
# ---------------------------------------------------------------------------

_GR_HEADER = ["id", "concepto", "categoria"]


def get_gastos_recurrentes():
    wb = _get_wb()
    ws = _ensure_sheet(wb, SHEET_GASTOS_RECURRENTES, _GR_HEADER)
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            result.append({"id": row[0], "concepto": row[1] or "", "categoria": row[2] or ""})
    return result


def add_gasto_recurrente(concepto: str, categoria: str = ""):
    wb = _get_wb()
    ws = _ensure_sheet(wb, SHEET_GASTOS_RECURRENTES, _GR_HEADER)
    new_id = _next_id(ws)
    ws.append([new_id, concepto.strip(), categoria.strip()])
    _save_wb(wb)
    return new_id


def delete_gasto_recurrente(gid: int):
    wb = _get_wb()
    ws = _ensure_sheet(wb, SHEET_GASTOS_RECURRENTES, _GR_HEADER)
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] is not None and str(row[0]) == str(gid):
            ws.delete_rows(i)
            break
    _save_wb(wb)


# ---------------------------------------------------------------------------
# TAREAS PENDIENTES
# ---------------------------------------------------------------------------

def get_tareas():
    wb = _get_wb()
    ws = _ensure_sheet(wb, SHEET_TAREAS, ["id", "descripcion"])
    tareas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            tareas.append({"id": int(row[0]), "descripcion": str(row[1] or "")})
    return tareas


def add_tarea(descripcion: str):
    wb = _get_wb()
    ws = _ensure_sheet(wb, SHEET_TAREAS, ["id", "descripcion"])
    new_id = _next_id(ws)
    ws.append([new_id, descripcion.strip()])
    _save_wb(wb)
    return {"id": new_id, "descripcion": descripcion.strip()}


def delete_tarea(tarea_id: int):
    wb = _get_wb()
    ws = _ensure_sheet(wb, SHEET_TAREAS, ["id", "descripcion"])
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] is not None and int(row[0]) == tarea_id:
            ws.delete_rows(i)
            _save_wb(wb)
            return True
    return False


# ---------------------------------------------------------------------------
# FACTURAS
# ---------------------------------------------------------------------------

_FAC_HEADER = ["id", "fecha", "proveedor_id", "proveedor_nombre", "descripcion",
               "importe", "estado", "fecha_pago", "categoria", "numero_factura", "extraordinario", "archivo_pdf"]

def get_facturas(estado=None):
    wb = _get_wb()
    ws = _ensure_sheet(wb, SHEET_FACTURAS, _FAC_HEADER)
    result = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] is not None:
            f = dict(zip(_FAC_HEADER, row)) | {"row": i}
            f["importe"] = float(f["importe"]) if f["importe"] else 0.0
            fecha = f.get("fecha")
            if hasattr(fecha, "strftime"):
                f["fecha"] = fecha.strftime("%Y-%m-%d")
            fecha_pago = f.get("fecha_pago")
            if hasattr(fecha_pago, "strftime"):
                f["fecha_pago"] = fecha_pago.strftime("%Y-%m-%d")
            if estado is None or f["estado"] == estado:
                result.append(f)
    result.sort(key=lambda x: x.get("fecha") or "", reverse=True)
    return result


def save_factura(data: dict):
    wb = _get_wb()
    ws = _ensure_sheet(wb, SHEET_FACTURAS, _FAC_HEADER)
    fid = data.get("id")
    # Convertir fechas
    for campo in ("fecha", "fecha_pago"):
        v = data.get(campo)
        if v and isinstance(v, str) and v.strip():
            try:
                data[campo] = datetime.strptime(v, "%Y-%m-%d").date()
            except ValueError:
                data[campo] = None
        elif not v:
            data[campo] = None

    if fid:
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] is not None and str(row[0]) == str(fid):
                for j, col in enumerate(_FAC_HEADER, start=1):
                    ws.cell(row=i, column=j).value = data.get(col)
                _save_wb(wb)
                return int(fid)
    new_id = _next_id(ws)
    data["id"] = new_id
    ws.append([data.get(col) for col in _FAC_HEADER])
    _save_wb(wb)

    # Auto-crear gasto si la descripción coincide con un gasto recurrente configurado
    desc = (data.get("descripcion") or "").strip()
    fecha = data.get("fecha")
    if desc and fecha:
        recurrentes = get_gastos_recurrentes()
        for gr in recurrentes:
            if gr["concepto"].strip().upper() == desc.upper():
                periodo = str(fecha)[:7]
                save_gasto(periodo, gr["concepto"], float(data.get("importe") or 0), "FIJO")
                break

    return new_id


def pagar_factura(fid: int, fecha_pago: str):
    """Marca factura como pagada y registra en caja diaria."""
    wb = _get_wb()
    ws = _ensure_sheet(wb, SHEET_FACTURAS, _FAC_HEADER)
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] is not None and str(row[0]) == str(fid):
            f = dict(zip(_FAC_HEADER, row))
            # Actualizar estado
            ws.cell(row=i, column=_FAC_HEADER.index("estado") + 1).value = "PAGADA"
            try:
                fecha_dt = datetime.strptime(fecha_pago, "%Y-%m-%d").date()
            except (ValueError, TypeError):
                fecha_dt = date.today()
            ws.cell(row=i, column=_FAC_HEADER.index("fecha_pago") + 1).value = fecha_dt
            _save_wb(wb)
            # Registrar en caja
            proveedor = f.get("proveedor_nombre") or f"Proveedor ID {f.get('proveedor_id')}"
            nro = f.get("numero_factura") or ""
            desc = f"Pago factura {nro} - {proveedor} - {f.get('descripcion','')}"
            importe = float(f["importe"]) if f.get("importe") else 0.0
            save_movimiento(
                fecha=fecha_pago,
                descripcion=desc[:200],
                tipo="SALIDA",
                categoria=f.get("categoria") or "Proveedor",
                importe=importe,
            )
            # Registrar como gasto mensual, excepto si ya fue creado al registrar la factura
            # recurrente, o si es un gasto de Fondo de Reserva (ya está en GASTOS_MENSUALES como VARIABLE_FR)
            periodo = fecha_dt.strftime("%Y-%m")
            concepto = f.get("descripcion") or proveedor
            recurrentes = get_gastos_recurrentes()
            es_recurrente = any(
                gr["concepto"].strip().upper() == concepto.strip().upper()
                for gr in recurrentes
            )
            es_fondo_reserva = (f.get("categoria") or "").upper() == "FONDO_RESERVA"
            if not es_recurrente and not es_fondo_reserva:
                save_gasto(periodo, concepto[:100], importe, f.get("categoria") or "FIJO")
            return True
    return False


def factura_en_liquidacion(fid) -> bool:
    """Retorna True si la factura ya fue incluida en una liquidación generada."""
    for f in get_facturas():
        if str(f["id"]) != str(fid):
            continue
        if f.get("estado") != "PAGADA":
            return False
        fecha_pago = f.get("fecha_pago") or ""
        if len(fecha_pago) >= 7:
            periodo_gasto = fecha_pago[:7]
            periodo_liq = _next_periodo(periodo_gasto)
            return liquidacion_existe(periodo_liq)
    return False


def get_facturas_extraordinarias_periodo(periodo: str):
    """Returns extraordinary facturas paid in the given period."""
    result = []
    for f in get_facturas(estado="PAGADA"):
        if not f.get("extraordinario"):
            continue
        fecha_pago = str(f.get("fecha_pago") or "")
        if fecha_pago[:7] == periodo:
            result.append(f)
    return result


def delete_factura(fid):
    if factura_en_liquidacion(fid):
        return False  # Bloqueada, ya está en una liquidación
    wb = _get_wb()
    ws = _ensure_sheet(wb, SHEET_FACTURAS, _FAC_HEADER)
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] is not None and str(row[0]) == str(fid):
            ws.delete_rows(i)
            break
    _save_wb(wb)
    return True


def set_factura_archivo_pdf(fid: int, ruta_relativa: str):
    """Guarda la ruta relativa del PDF adjunto en la columna archivo_pdf."""
    wb = _get_wb()
    ws = _ensure_sheet(wb, SHEET_FACTURAS, _FAC_HEADER)
    col_idx = _FAC_HEADER.index("archivo_pdf") + 1
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] is not None and str(row[0]) == str(fid):
            ws.cell(row=i, column=col_idx).value = ruta_relativa
            _save_wb(wb)
            return True
    return False


# ---------------------------------------------------------------------------
# PEDIDOS DE PRESUPUESTO
# ---------------------------------------------------------------------------

_PED_HEADER = ["id", "fecha", "descripcion", "categoria", "estado", "proveedor_elegido", "notas"]

def get_pedidos():
    wb = _get_wb()
    ws = _ensure_sheet(wb, SHEET_PEDIDOS, _PED_HEADER)
    result = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] is not None:
            p = dict(zip(_PED_HEADER, row)) | {"row": i}
            fecha = p.get("fecha")
            if hasattr(fecha, "strftime"):
                p["fecha"] = fecha.strftime("%Y-%m-%d")
            result.append(p)
    result.sort(key=lambda x: x.get("fecha") or "", reverse=True)
    return result


def save_pedido(data: dict):
    wb = _get_wb()
    ws = _ensure_sheet(wb, SHEET_PEDIDOS, _PED_HEADER)
    v = data.get("fecha")
    if v and isinstance(v, str):
        try:
            data["fecha"] = datetime.strptime(v, "%Y-%m-%d").date()
        except ValueError:
            data["fecha"] = date.today()
    pid = data.get("id")
    if pid:
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] is not None and str(row[0]) == str(pid):
                for j, col in enumerate(_PED_HEADER, start=1):
                    ws.cell(row=i, column=j).value = data.get(col, "")
                _save_wb(wb)
                return int(pid)
    new_id = _next_id(ws)
    data["id"] = new_id
    ws.append([data.get(col, "") for col in _PED_HEADER])
    _save_wb(wb)
    return new_id


def delete_pedido(pid):
    wb = _get_wb()
    ws = _ensure_sheet(wb, SHEET_PEDIDOS, _PED_HEADER)
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] is not None and str(row[0]) == str(pid):
            ws.delete_rows(i)
            break
    # También eliminar los presupuestos asociados
    ws2 = _ensure_sheet(wb, SHEET_PRESUPUESTOS, _PRES_HEADER)
    rows_del = [i for i, row in enumerate(ws2.iter_rows(min_row=2, values_only=True), start=2)
                if row[1] is not None and str(row[1]) == str(pid)]
    for r in reversed(rows_del):
        ws2.delete_rows(r)
    _save_wb(wb)


# ---------------------------------------------------------------------------
# PRESUPUESTOS (cotizaciones por pedido)
# ---------------------------------------------------------------------------

_PRES_HEADER = ["id", "pedido_id", "proveedor_id", "proveedor_nombre", "fecha",
                "importe", "notas", "seleccionado"]

def get_presupuestos(pedido_id=None):
    wb = _get_wb()
    ws = _ensure_sheet(wb, SHEET_PRESUPUESTOS, _PRES_HEADER)
    result = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] is not None:
            p = dict(zip(_PRES_HEADER, row)) | {"row": i}
            p["importe"] = float(p["importe"]) if p["importe"] else 0.0
            p["seleccionado"] = int(p["seleccionado"]) if p["seleccionado"] else 0
            fecha = p.get("fecha")
            if hasattr(fecha, "strftime"):
                p["fecha"] = fecha.strftime("%Y-%m-%d")
            if pedido_id is None or str(p["pedido_id"]) == str(pedido_id):
                result.append(p)
    return result


def save_presupuesto(data: dict):
    wb = _get_wb()
    ws = _ensure_sheet(wb, SHEET_PRESUPUESTOS, _PRES_HEADER)
    v = data.get("fecha")
    if v and isinstance(v, str):
        try:
            data["fecha"] = datetime.strptime(v, "%Y-%m-%d").date()
        except ValueError:
            data["fecha"] = date.today()
    pid = data.get("id")
    if pid:
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] is not None and str(row[0]) == str(pid):
                for j, col in enumerate(_PRES_HEADER, start=1):
                    ws.cell(row=i, column=j).value = data.get(col)
                _save_wb(wb)
                return int(pid)
    new_id = _next_id(ws)
    data["id"] = new_id
    ws.append([data.get(col) for col in _PRES_HEADER])
    _save_wb(wb)
    return new_id


def seleccionar_presupuesto(pres_id: int, pedido_id: int):
    """Marca un presupuesto como seleccionado y desmarca los demás del mismo pedido."""
    wb = _get_wb()
    ws = _ensure_sheet(wb, SHEET_PRESUPUESTOS, _PRES_HEADER)
    proveedor_elegido = ""
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] is not None and str(row[1]) == str(pedido_id):
            es_este = str(row[0]) == str(pres_id)
            ws.cell(row=i, column=_PRES_HEADER.index("seleccionado") + 1).value = 1 if es_este else 0
            if es_este:
                proveedor_elegido = str(row[3] or "")
    # Actualizar pedido con proveedor elegido
    ws2 = _ensure_sheet(wb, SHEET_PEDIDOS, _PED_HEADER)
    for i, row in enumerate(ws2.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] is not None and str(row[0]) == str(pedido_id):
            ws2.cell(row=i, column=_PED_HEADER.index("proveedor_elegido") + 1).value = proveedor_elegido
            ws2.cell(row=i, column=_PED_HEADER.index("estado") + 1).value = "ADJUDICADO"
            break
    _save_wb(wb)


def delete_presupuesto(pres_id):
    wb = _get_wb()
    ws = _ensure_sheet(wb, SHEET_PRESUPUESTOS, _PRES_HEADER)
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] is not None and str(row[0]) == str(pres_id):
            ws.delete_rows(i)
            break
    _save_wb(wb)


def get_saldo_caja():
    """Saldo acumulado total (banco): saldo_inicial + entradas - salidas."""
    cfg = get_config()
    saldo_inicial = float(cfg.get("saldo_inicial_caja", 0) or 0)
    wb = _get_wb()
    ws = _ensure_sheet(wb, SHEET_CAJA, ["fecha", "descripcion", "tipo", "categoria", "importe"])
    entradas = salidas = 0.0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        importe = float(row[4]) if row[4] else 0.0
        if row[2] == "ENTRADA":
            entradas += importe
        elif row[2] == "SALIDA":
            salidas += importe
    saldo = round(saldo_inicial + entradas - salidas, 2)
    return saldo, round(entradas, 2), round(salidas, 2)


def get_movimientos_periodo(periodo: str):
    """Entradas y salidas filtradas solo por el período 'YYYY-MM'."""
    wb = _get_wb()
    ws = _ensure_sheet(wb, SHEET_CAJA, ["fecha", "descripcion", "tipo", "categoria", "importe"])
    entradas = salidas = 0.0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        fecha = row[0]
        if hasattr(fecha, "strftime"):
            fecha_str = fecha.strftime("%Y-%m")
        else:
            fecha_str = str(fecha)[:7]
        if fecha_str != periodo:
            continue
        importe = float(row[4]) if row[4] else 0.0
        if row[2] == "ENTRADA":
            entradas += importe
        elif row[2] == "SALIDA":
            salidas += importe
    return round(entradas, 2), round(salidas, 2)


def get_fondo_reserva():
    """Retorna el acumulado del fondo de reserva: suma de todos los gastos tipo FONDO_RESERVA."""
    wb = _get_wb()
    ws = wb[SHEET_GASTOS]
    total = 0.0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[3] == "FONDO_RESERVA" and row[2]:
            total += float(row[2])
    return round(total, 2)


def ensure_fondo_reserva_gasto(periodo: str):
    """
    Agrega automáticamente el gasto de Fondo de Reserva para el período si no existe.
    No hace nada si el monto mensual no está configurado o ya existe el gasto.
    """
    cfg = get_config()
    monto = float(cfg.get("fondo_reserva_mensual", 0) or 0)
    if monto <= 0:
        return
    gastos = get_gastos(periodo)
    if any(g["tipo"] == "FONDO_RESERVA" for g in gastos):
        return  # Ya existe para este período
    save_gasto(periodo, "Fondo de Reserva", monto, "FONDO_RESERVA")


def get_historial_unidades():
    """Retorna historial de liquidaciones de todas las unidades, agrupado por unidad."""
    wb = _get_wb()
    historial = {}  # {numero: [filas]}
    for sheet_name in sorted(wb.sheetnames):
        if not sheet_name.startswith(LIQPREFIX):
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 10 or not row[0] or not row[1]:
                continue
            numero = str(row[1])
            fecha_p = row[14] if len(row) > 14 else None
            if hasattr(fecha_p, "strftime"):
                fecha_p = fecha_p.strftime("%Y-%m-%d")
            entry = {
                "periodo":        str(row[0]),
                "descripcion":    row[2] or "",
                "propietario":    row[3] or "",
                "inquilino":      row[4] or "",
                "expensas":       float(row[6]) if row[6] else 0.0,
                "deuda_anterior": float(row[7]) if row[7] else 0.0,
                "interes":        float(row[8]) if row[8] else 0.0,
                "total_a_pagar":  float(row[9]) if row[9] else 0.0,
                "monto_pagado":   float(row[11]) if len(row) > 11 and row[11] else 0.0,
                "tipo_pago":      row[12] if len(row) > 12 and row[12] else "PENDIENTE",
                "saldo_pendiente":float(row[13]) if len(row) > 13 and row[13] else 0.0,
                "fecha_pago":     fecha_p or "",
            }
            historial.setdefault(numero, []).append(entry)
    # Ordenar cada unidad por período
    for num in historial:
        historial[num].sort(key=lambda x: x["periodo"])
    return historial


def get_apertura():
    """Retorna saldo inicial de caja y deudas iniciales por unidad."""
    cfg = get_config()
    saldo_inicial = float(cfg.get("saldo_inicial_caja", 0) or 0)
    unidades = get_unidades()
    return saldo_inicial, unidades


def save_apertura(saldo_inicial: float, deudas: dict):
    """Guarda saldo inicial de caja y deuda_inicial por unidad."""
    # Guardar saldo inicial en config
    cfg_data = {"saldo_inicial_caja": str(saldo_inicial)}
    save_config(cfg_data)
    # Guardar deuda_inicial en cada unidad
    wb = _get_wb()
    ws = wb[SHEET_UNIDADES]
    header = [c.value for c in ws[1]]
    if "deuda_inicial" not in header:
        header.append("deuda_inicial")
        ws.cell(row=1, column=len(header)).value = "deuda_inicial"
    col_deuda = header.index("deuda_inicial") + 1
    col_num = header.index("numero") + 1
    for row in ws.iter_rows(min_row=2):
        num = str(row[col_num - 1].value)
        if num in deudas:
            row[col_deuda - 1].value = deudas[num]
    _save_wb(wb)


def reset_datos_operativos():
    """Borra liquidaciones, gastos, caja y facturas. Preserva unidades y proveedores."""
    wb = _get_wb()
    for sheet in [SHEET_GASTOS, SHEET_CAJA, SHEET_FACTURAS]:
        ws = _ensure_sheet(wb, sheet, [])
        for r in range(ws.max_row, 1, -1):
            ws.delete_rows(r)
    # Borrar hojas de liquidaciones (excluyendo la hoja de estados)
    for name in [n for n in wb.sheetnames if n.startswith(LIQPREFIX) and n != SHEET_LIQ_ESTADOS]:
        del wb[name]
    # Limpiar hoja de estados
    ws_est = _ensure_sheet(wb, SHEET_LIQ_ESTADOS, _LIQ_EST_HEADER)
    for r in range(ws_est.max_row, 1, -1):
        ws_est.delete_rows(r)
    _save_wb(wb)


def get_estado_plano(periodo: str):
    """Retorna dict {unidad: {datos}} para el plano del edificio."""
    liq = {str(r["unidad"]): r for r in get_liquidacion(periodo)}
    unidades = get_unidades()
    resultado = []
    for u in unidades:
        num = str(u["numero"])
        liq_row = liq.get(num)
        tipo_pago = liq_row.get("tipo_pago", "PENDIENTE") if liq_row else "PENDIENTE"
        resultado.append({
            "numero": num,
            "descripcion": u.get("descripcion", ""),
            "propietario": u.get("propietario", ""),
            "inquilino": u.get("inquilino", ""),
            "piso": u.get("piso", "") or "",
            "activo": u.get("activo", 1),
            "deuda_inicial": u.get("deuda_inicial", 0) or 0,
            "tiene_liquidacion": liq_row is not None,
            "pagado": 1 if tipo_pago == "TOTAL" else (2 if tipo_pago == "PARCIAL" else 0),
            "tipo_pago": tipo_pago,
            "total_a_pagar": liq_row["total_a_pagar"] if liq_row else 0,
            "monto_pagado": liq_row.get("monto_pagado", 0) if liq_row else 0,
            "saldo_pendiente": liq_row.get("saldo_pendiente", 0) if liq_row else 0,
            "expensas": liq_row["expensas"] if liq_row else 0,
            "deuda_anterior": liq_row["deuda_anterior"] if liq_row else 0,
            "interes": liq_row["interes"] if liq_row else 0,
            "pct": liq_row["pct_aplicado"] if liq_row else 0,
        })
    return resultado
